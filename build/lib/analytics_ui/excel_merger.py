import tkinter as tk
from tkinter import filedialog, ttk, messagebox
import pandas as pd
import os
import sys
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import PatternFill, Border, Side, Alignment
import numpy as np
import logging

def resource_path(relative_path):
    """Получает абсолютный путь к ресурсу, работает для dev, PyInstaller и pip install"""
    try:
        # PyInstaller создает временную папку и хранит путь в _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        # Normal python script or installed package
        base_path = os.path.dirname(os.path.abspath(__file__))
    
    path = os.path.join(base_path, relative_path)
    # print(f"Debug: Looking for resource at {path}") # Uncomment for debugging
    return path

def get_column_letter(col_idx):
    """Преобразует числовой индекс столбца в буквенное обозначение Excel"""
    result = ""
    while col_idx > 0:
        col_idx, remainder = divmod(col_idx - 1, 26)
        result = chr(65 + remainder) + result
    return result

def apply_conditional_formatting(file_path):
    """Применяет условное форматирование к файлу Excel"""
    try:
        print(f"Начинаем применять форматирование к файлу: {file_path}")
        
        # Загружаем файл
        wb = load_workbook(filename=file_path)
        ws = wb.active
        
        # Вставляем пустую строку сверху для диапазонов
        ws.insert_rows(1)
        
        # Закрепляем заголовки (теперь они на 2 строке) и первый столбец
        ws.freeze_panes = 'B3'
        
        # Получаем максимальное количество строк и столбцов
        max_row = ws.max_row
        max_col = ws.max_column
        print(f"Размер таблицы: {max_row} строк, {max_col} столбцов")
        
        # Получаем заголовки столбцов (теперь на 2 строке)
        headers = [ws.cell(row=2, column=i).value for i in range(1, max_col + 1)]
        print(f"Заголовки столбцов: {headers}")
        
        # Создаем правило цветовой шкалы с использованием стандартных цветов Excel
        color_scale_rule = ColorScaleRule(
            start_type='num',
            start_value=0.000001,  # Минимальное значение, близкое к нулю
            start_color='FFFF0000',  # Красный для низких значений
            mid_type='percentile',
            mid_value=50,
            mid_color='FFFFFF00',    # Желтый для средних значений
            end_type='max',
            end_color='FF92D050'     # Зеленый для высоких значений
        )
        
        # Читаем файл с правилами для получения информации об узлах измерения
        rules_file = resource_path("Правила названия столбцов.xlsx")
        rules_df = pd.read_excel(rules_file, engine='openpyxl')
        
        # Создаем словарь соответствия столбцов и узлов измерения, а также диапазонов
        column_to_node = {}
        param_ranges = {}
        
        for _, row in rules_df.iterrows():
            if len(row) >= 4:
                new_name = str(row.iloc[2]).strip()
                node_name = str(row.iloc[3]).strip()
                if new_name and node_name:
                    column_to_node[new_name] = node_name
                
                # Попытка получить диапазоны и единицы измерения
                if len(row) >= 7 and new_name:
                    try:
                        min_val = row.iloc[5]
                        max_val = row.iloc[6]
                        units = str(row.iloc[7]).strip() if len(row) >= 8 and pd.notna(row.iloc[7]) else ""
                        
                        # Если единицы измерения не заданы явно, определяем по имени параметра
                        if not units:
                            # Проверяем 4-й столбец (имя параметра), если он существует
                            param_name = str(row.iloc[4]).strip() if len(row) >= 5 else ""
                            check_name = param_name.lower() if param_name else new_name.lower()
                            
                            if "перепад давления" in check_name:
                                units = "кгс/см2"
                            elif "расход" in check_name:
                                units = "тыс. м3/ч"
                            elif "температура" in check_name:
                                units = "°C"
                        
                        if pd.notna(min_val) and pd.notna(max_val):
                            range_str = f"({min_val} ... {max_val} {units})".strip()
                            param_ranges[new_name] = range_str
                    except Exception:
                        pass
        
        # Группируем столбцы по узлам измерения
        current_node = None
        node_start_col = None
        
        # Создаем стили границ
        thin_border = Side(style='thin')
        thick_border = Side(style='thick')
        
        # Создаем стиль выравнивания по центру
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Применяем форматирование и границы
        formatted_columns = 0
        
        # Словарь для отслеживания пар Параметр-Стрелка для объединения
        merge_candidates = {} # {param_name: {'start_col': idx, 'end_col': idx}}
        
        for col_idx in range(1, max_col + 1):
            try:
                header = headers[col_idx - 1]
                base_header = header.split(' ⚠')[0] if header and ' ⚠' in header else header
                
                # Если это заголовок с диапазоном, готовим объединение
                if base_header in param_ranges:
                    if base_header not in merge_candidates:
                        merge_candidates[base_header] = {'start': col_idx, 'end': col_idx}
                    else:
                        merge_candidates[base_header]['end'] = col_idx

                # Определяем узел измерения для текущего столбца
                node = column_to_node.get(base_header)
                
                # Обработка границ для группировки по узлам
                if node:
                    if node != current_node:
                        # Закрываем предыдущую группу
                        if current_node and node_start_col:
                            # Устанавливаем правую границу для предыдущей группы
                            for row in range(1, max_row + 1):
                                cell = ws.cell(row=row, column=col_idx-1)
                                current_border = cell.border
                                cell.border = Border(
                                    left=current_border.left,
                                    right=thick_border,
                                    top=current_border.top,
                                    bottom=current_border.bottom
                                )
                        
                        # Начинаем новую группу
                        current_node = node
                        node_start_col = col_idx
                        
                        # Устанавливаем левую границу для новой группы
                        for row in range(1, max_row + 1):
                            cell = ws.cell(row=row, column=col_idx)
                            current_border = cell.border
                            cell.border = Border(
                                left=thick_border,
                                right=current_border.right,
                                top=current_border.top,
                                bottom=current_border.bottom
                            )
                
                # Применяем условное форматирование (данные теперь с 3 строки)
                if header != "Время" and not header.endswith('⚠'):
                    col_letter = get_column_letter(col_idx)
                    
                    # Создаем список диапазонов для форматирования, исключая ячейки с нулевыми значениями
                    ranges_to_format = []
                    current_range_start = None
                    
                    for row in range(3, max_row + 1):
                        cell = ws[f'{col_letter}{row}']
                        cell_value = cell.value
                        
                        try:
                            if cell_value is not None:
                                numeric_value = float(str(cell_value).replace(',', '.').strip())
                                if numeric_value != 0:
                                    if current_range_start is None:
                                        current_range_start = row
                                else:
                                    if current_range_start is not None:
                                        range_str = f'{col_letter}{current_range_start}:{col_letter}{row-1}'
                                        ranges_to_format.append(range_str)
                                        current_range_start = None
                        except (ValueError, TypeError):
                            if current_range_start is not None:
                                range_str = f'{col_letter}{current_range_start}:{col_letter}{row-1}'
                                ranges_to_format.append(range_str)
                                current_range_start = None
                    
                    # Добавляем последний диапазон, если он есть
                    if current_range_start is not None:
                        range_str = f'{col_letter}{current_range_start}:{col_letter}{max_row}'
                        ranges_to_format.append(range_str)
                    
                    # Применяем форматирование к каждому диапазону
                    for range_str in ranges_to_format:
                        print(f"Применяем форматирование к диапазону: {range_str}")
                        ws.conditional_formatting.add(range_str, color_scale_rule)
                    
                    if ranges_to_format:
                        print(f"Форматирование успешно применено к столбцу {header}")
                        formatted_columns += 1
                    else:
                        print(f"В столбце {header} нет ненулевых значений для форматирования")
                
                # Применяем выравнивание по центру ко всем ячейкам в столбце (включая новые заголовки)
                for row in range(1, max_row + 1):
                    ws.cell(row=row, column=col_idx).alignment = center_alignment
                
            except Exception as col_error:
                print(f"Ошибка при форматировании столбца {get_column_letter(col_idx)}: {str(col_error)}")
                continue
        
        # Применяем объединение и тексты диапазонов
        for param, coords in merge_candidates.items():
            if param in param_ranges:
                start_col = coords['start']
                end_col = coords['end']
                # Объединяем ячейки
                ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
                # Пишем текст
                cell = ws.cell(row=1, column=start_col)
                cell.value = param_ranges[param]
                cell.alignment = center_alignment

        # Автоподбор ширины столбцов
        for col_idx in range(1, max_col + 1):
            col_letter = get_column_letter(col_idx)
            max_length = 0
            
            # Проверяем заголовки (строки 1 и 2)
            for row in [1, 2]:
                cell = ws.cell(row=row, column=col_idx)
                if cell.value:
                    # Если ячейка объединена, длину делим на количество столбцов (грубо)
                    is_merged = False
                    for merged_range in ws.merged_cells.ranges:
                        if cell.coordinate in merged_range:
                            # Пропускаем расчет по объединенным ячейкам для простоты, 
                            # или можно брать часть длины.
                            # Здесь лучше ориентироваться на заголовок (стр 2) и данные
                            if row == 1: is_merged = True
                            break
                    
                    if not is_merged:
                         max_length = max(max_length, len(str(cell.value)))

            # Проверяем первые 100 строк данных для скорости
            for row in range(3, min(max_row + 1, 103)):
                cell = ws.cell(row=row, column=col_idx)
                if cell.value:
                    try:
                         max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
            
            # Устанавливаем ширину с небольшим запасом
            adjusted_width = (max_length + 2)
            # Ограничиваем разумными пределами
            adjusted_width = min(max(adjusted_width, 10), 50) 
            ws.column_dimensions[col_letter].width = adjusted_width

        # Закрываем последнюю группу
        if current_node and node_start_col:
            for row in range(1, max_row + 1):
                cell = ws.cell(row=row, column=max_col)
                current_border = cell.border
                cell.border = Border(
                    left=current_border.left,
                    right=thick_border,
                    top=current_border.top,
                    bottom=current_border.bottom
                )
        
        # Добавляем верхнюю и нижнюю границы для всех ячеек
        for col in range(1, max_col + 1):
            # Верхняя граница (теперь строка 1)
            cell = ws.cell(row=1, column=col)
            current_border = cell.border
            cell.border = Border(
                left=current_border.left,
                right=current_border.right,
                top=thick_border,
                bottom=current_border.bottom # Исправляем на обычную или сохраняем логику
            )
             # Граница под заголовками (строка 2)
            cell = ws.cell(row=2, column=col)
            current_border = cell.border
            # Здесь можно добавить линию разделитель
            cell.border = Border(
                 left=current_border.left,
                 right=current_border.right,
                 top=current_border.top, # тонкая или толстая
                 bottom=thin_border
            )

            # Нижняя граница для последней строки
            cell = ws.cell(row=max_row, column=col)
            current_border = cell.border
            cell.border = Border(
                left=current_border.left,
                right=current_border.right,
                top=current_border.top,
                bottom=thick_border
            )
        
        print(f"Всего отформатировано столбцов: {formatted_columns}")
        
        # Сохраняем изменения
        print("Сохраняем изменения в файл...")
        try:
            # Сначала сохраняем во временный файл
            temp_file = file_path.replace('.xlsx', '_temp.xlsx')
            wb.save(temp_file)
            wb.close()
            
            # Если сохранение прошло успешно, заменяем исходный файл
            import os
            if os.path.exists(file_path):
                os.remove(file_path)
            os.rename(temp_file, file_path)
            
            print("Форматирование успешно применено и файл сохранен")
            return True
            
        except Exception as save_error:
            print(f"Ошибка при сохранении файла: {str(save_error)}")
            return False
        
    except Exception as e:
        error_msg = f"Ошибка при применении условного форматирования: {str(e)}"
        print(error_msg)
        print(f"Тип ошибки: {type(e).__name__}")
        import traceback
        print(f"Полный стек ошибки:\n{traceback.format_exc()}")
        return False
        
    except Exception as e:
        error_msg = f"Ошибка при применении условного форматирования: {str(e)}"
        print(error_msg)
        print(f"Тип ошибки: {type(e).__name__}")
        import traceback
        print(f"Полный стек ошибки:\n{traceback.format_exc()}")
        return False

def add_arrow_columns(df, rules_file):
    """Добавляет столбцы со стрелками для значений вне диапазона min-max"""
    try:
        # Читаем файл с правилами
        rules_df = pd.read_excel(rules_file, engine='openpyxl')
        
        # Создаем словари для хранения min/max значений параметров и соответствия узлов измерения
        param_limits = {}
        param_to_node = {}  # Словарь для хранения соответствия параметр -> узел измерения
        
        # Заполняем словари из файла правил
        for _, row in rules_df.iterrows():
            if len(row) >= 7:  # Проверяем наличие столбцов min и max
                new_name = str(row.iloc[2]).strip()
                node_name = str(row.iloc[3]).strip()  # Название узла измерения
                min_val = row.iloc[5]  # Предполагаем, что min в 6-м столбце
                max_val = row.iloc[6]  # Предполагаем, что max в 7-м столбце
                
                # Сохраняем соответствие параметра и узла измерения
                if new_name and node_name:
                    param_to_node[new_name] = node_name
                
                # Проверяем, что значения min и max заданы
                if pd.notna(min_val) and pd.notna(max_val) and new_name:
                    try:
                        min_val = float(min_val)
                        max_val = float(max_val)
                        param_limits[new_name] = {'min': min_val, 'max': max_val}
                    except (ValueError, TypeError):
                        print(f"Пропущены некорректные значения min/max для {new_name}")
                        continue
        
        # Добавляем столбцы со стрелками
        for col in df.columns:
            if col in param_limits:
                limits = param_limits[col]
                arrow_col_name = f"{col} ⚠"
                
                # Создаем столбец со стрелками
                df[arrow_col_name] = ''
                
                # Заполняем стрелки на основе условий, пропуская нулевые значения
                numeric_values = pd.to_numeric(df[col], errors='coerce')
                df.loc[(numeric_values < limits['min']) & (numeric_values != 0), arrow_col_name] = '↓'
                df.loc[(numeric_values > limits['max']) & (numeric_values != 0), arrow_col_name] = '↑'
                
                # Перемещаем столбец со стрелками сразу после исходного столбца
                cols = list(df.columns)
                current_idx = cols.index(arrow_col_name)
                target_idx = cols.index(col) + 1
                cols.insert(target_idx, cols.pop(current_idx))
                df = df[cols]
        
        return df, param_to_node
        
    except Exception as e:
        print(f"Ошибка при добавлении столбцов со стрелками: {str(e)}")
        return df, {}

class ExcelMerger:
    def __init__(self, root):
        self.root = root
        self.root.title("Объединение Excel файлов")
        self.root.geometry("1200x600")
        
        # Список для хранения путей к файлам
        self.files = []
        
        # Словари для хранения чекбоксов
        self.parameter_vars = {}
        self.node_vars = {}  # Новый словарь для чекбоксов узлов измерения
        
        # Создание элементов интерфейса
        self.create_widgets()
        
    def create_widgets(self):
        # Создаем фрейм для левой части (файлы)
        left_frame = ttk.Frame(self.root)
        left_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Кнопка для добавления файлов
        add_button = ttk.Button(left_frame, text="Добавить файлы", command=self.add_files)
        add_button.pack(pady=10)

        # Кнопка для настройки диапазонов
        ranges_button = ttk.Button(left_frame, text="Настройка диапазонов", command=self.open_range_editor)
        ranges_button.pack(pady=5)
        
        # Список файлов
        self.files_frame = ttk.LabelFrame(left_frame, text="Выбранные файлы")
        self.files_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        self.files_listbox = tk.Listbox(self.files_frame)
        self.files_listbox.pack(fill=tk.BOTH, expand=True)
        
        # Кнопка удаления выбранного файла
        remove_button = ttk.Button(left_frame, text="Удалить выбранный", command=self.remove_file)
        remove_button.pack(pady=5)
        
        # Фрейм для временного интервала
        time_frame = ttk.LabelFrame(left_frame, text="Временной интервал")
        time_frame.pack(fill=tk.X, padx=5, pady=5)
        
        # Начальная дата
        start_frame = ttk.Frame(time_frame)
        start_frame.pack(fill=tk.X, padx=5, pady=2)
        ttk.Label(start_frame, text="Начало:").pack(side=tk.LEFT)
        self.start_date = ttk.Entry(start_frame, width=10)
        self.start_date.pack(side=tk.LEFT, padx=5)
        self.start_time = ttk.Entry(start_frame, width=8)
        self.start_time.pack(side=tk.LEFT)
        
        # Конечная дата
        end_frame = ttk.Frame(time_frame)
        end_frame.pack(fill=tk.X, padx=5, pady=2)
        ttk.Label(end_frame, text="Конец:").pack(side=tk.LEFT)
        self.end_date = ttk.Entry(end_frame, width=10)
        self.end_date.pack(side=tk.LEFT, padx=5)
        self.end_time = ttk.Entry(end_frame, width=8)
        self.end_time.pack(side=tk.LEFT)
        
        # Подсказка формата
        ttk.Label(time_frame, text="Формат: ГГГГ-ММ-ДД ЧЧ:ММ:СС", font=("Arial", 8)).pack(pady=2)
        
        # Кнопка для установки полного диапазона
        set_range_button = ttk.Button(time_frame, text="Установить полный диапазон", command=self.set_full_time_range)
        set_range_button.pack(pady=5)
        
        # Кнопка для очистки дат
        clear_dates_button = ttk.Button(time_frame, text="Очистить даты", command=self.clear_dates)
        clear_dates_button.pack(pady=5)
        
        # Кнопка объединения
        merge_button = ttk.Button(left_frame, text="Объединить файлы", command=self.merge_files)
        merge_button.pack(pady=10)
        
        # Создаем центральный фрейм для параметров
        center_frame = ttk.Frame(self.root)
        center_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Фрейм для параметров
        self.parameters_frame = ttk.LabelFrame(center_frame, text="Выбор параметров")
        self.parameters_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Создаем канвас и скроллбар для прокрутки параметров
        canvas = tk.Canvas(self.parameters_frame)
        scrollbar = ttk.Scrollbar(self.parameters_frame, orient="vertical", command=canvas.yview)
        self.scrollable_frame = ttk.Frame(canvas)

        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Размещаем элементы
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Кнопки выбора параметров
        select_frame = ttk.Frame(center_frame)
        select_frame.pack(fill=tk.X, padx=5, pady=5)
        
        ttk.Button(select_frame, text="Выбрать все", command=self.select_all_parameters).pack(side=tk.LEFT, padx=5)
        ttk.Button(select_frame, text="Снять все", command=self.deselect_all_parameters).pack(side=tk.LEFT, padx=5)
        
        # Создаем фрейм для правой части (узлы измерения)
        right_frame = ttk.Frame(self.root)
        right_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Фрейм для узлов измерения
        self.nodes_frame = ttk.LabelFrame(right_frame, text="Загруженные узлы измерения")
        self.nodes_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Создаем канвас и скроллбар для прокрутки узлов
        nodes_canvas = tk.Canvas(self.nodes_frame)
        nodes_scrollbar = ttk.Scrollbar(self.nodes_frame, orient="vertical", command=nodes_canvas.yview)
        self.nodes_scrollable_frame = ttk.Frame(nodes_canvas)

        self.nodes_scrollable_frame.bind(
            "<Configure>",
            lambda e: nodes_canvas.configure(scrollregion=nodes_canvas.bbox("all"))
        )

        nodes_canvas.create_window((0, 0), window=self.nodes_scrollable_frame, anchor="nw")
        nodes_canvas.configure(yscrollcommand=nodes_scrollbar.set)
        
        # Размещаем элементы
        nodes_canvas.pack(side="left", fill="both", expand=True)
        nodes_scrollbar.pack(side="right", fill="y")
        
        # Кнопки выбора узлов
        nodes_select_frame = ttk.Frame(right_frame)
        nodes_select_frame.pack(fill=tk.X, padx=5, pady=5)
        
        ttk.Button(nodes_select_frame, text="Выбрать все", command=self.select_all_nodes).pack(side=tk.LEFT, padx=5)
        ttk.Button(nodes_select_frame, text="Снять все", command=self.deselect_all_nodes).pack(side=tk.LEFT, padx=5)
        
        # Загружаем параметры из файла правил
        self.load_parameters()
        
        # Авторское право
        copyright_label = ttk.Label(right_frame, text="© Н.А. Галаков", font=("Arial", 8), foreground="gray")
        copyright_label.pack(side=tk.BOTTOM, anchor=tk.E, padx=5, pady=5)
        
    def load_parameters(self):
        """Загружает параметры из файла правил"""
        try:
            # Очищаем старые чекбоксы
            for widget in self.scrollable_frame.winfo_children():
                if isinstance(widget, ttk.Checkbutton):
                    widget.destroy()
            self.parameter_vars.clear()
            
            # Читаем файл с правилами
            rules_file = resource_path("Правила названия столбцов.xlsx")
            if not os.path.exists(rules_file):
                messagebox.showerror("Ошибка", "Файл с правилами названия столбцов не найден!")
                return
                
            # Читаем файл с правилами
            rules_df = pd.read_excel(rules_file, engine='openpyxl')
            
            # Получаем уникальные параметры из 5-го столбца
            parameters = rules_df.iloc[:, 4].unique()
            parameters = [str(p).strip() for p in parameters if pd.notna(p) and str(p).strip()]
            
            # Создаем чекбоксы для каждого параметра
            for param in parameters:
                var = tk.BooleanVar(value=True)
                self.parameter_vars[param] = var
                ttk.Checkbutton(self.scrollable_frame, text=param, variable=var).pack(anchor="w", padx=5, pady=2)
                
        except Exception as e:
            error_message = f"Ошибка при загрузке параметров: {str(e)}"
            print(error_message)
            messagebox.showerror("Ошибка", error_message)
            
    def select_all_parameters(self):
        """Выбирает все параметры"""
        for var in self.parameter_vars.values():
            var.set(True)
            
    def deselect_all_parameters(self):
        """Снимает выбор со всех параметров"""
        for var in self.parameter_vars.values():
            var.set(False)
            
    def add_files(self):
        files = filedialog.askopenfilenames(
            title="Выберите Excel файлы",
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        for file in files:
            if file not in self.files:
                self.files.append(file)
                self.files_listbox.insert(tk.END, os.path.basename(file))
                # Обновляем список узлов измерения при добавлении файла
                self.update_measurement_nodes()
                # Обновляем список параметров
                self.load_parameters()
                # Обновляем временной диапазон
                df = self.read_excel_file(file)
                self.update_time_range(df)
                
    def remove_file(self):
        selection = self.files_listbox.curselection()
        if selection:
            index = selection[0]
            self.files.pop(index)
            self.files_listbox.delete(index)
            # Обновляем список узлов измерения при удалении файла
            self.update_measurement_nodes()
            # Обновляем список параметров
            self.load_parameters()
            # Сбрасываем временной диапазон
            self.clear_dates()
            
    def select_all_nodes(self):
        """Выбирает все узлы измерения"""
        for var in self.node_vars.values():
            var.set(True)
            
    def deselect_all_nodes(self):
        """Снимает выбор со всех узлов измерения"""
        for var in self.node_vars.values():
            var.set(False)
            
    def update_measurement_nodes(self):
        """Обновляет список узлов измерения на основе загруженных файлов"""
        try:
            # Очищаем старые чекбоксы узлов
            for widget in self.nodes_scrollable_frame.winfo_children():
                if isinstance(widget, ttk.Checkbutton):
                    widget.destroy()
            self.node_vars.clear()
            
            # Читаем файл с правилами
            rules_file = resource_path("Правила названия столбцов.xlsx")
            if not os.path.exists(rules_file):
                messagebox.showerror("Ошибка", "Файл с правилами названия столбцов не найден!")
                return
                
            rules_df = pd.read_excel(rules_file, engine='openpyxl')
            
            # Множество для хранения уникальных узлов измерения
            measurement_nodes = set()
            
            # Для каждого файла находим соответствующие узлы измерения
            for file in self.files:
                filename = os.path.basename(file).lower()
                
                # Ищем соответствующие правила
                for _, row in rules_df.iterrows():
                    if len(row) >= 4:  # Проверяем наличие столбца с названием узла
                        file_pattern = str(row.iloc[0]).strip().lower()
                        node_name = str(row.iloc[3]).strip()  # Предполагаем, что название узла в 4-м столбце
                        
                        # Проверяем, что значения не являются NaN и не пустые
                        if (pd.notna(file_pattern) and pd.notna(node_name) 
                            and file_pattern and node_name 
                            and file_pattern in filename
                            and node_name.lower() != 'nan'):  # Добавляем проверку на 'nan'
                            measurement_nodes.add(node_name)
            
            # Добавляем найденные узлы и создаем чекбоксы
            for node in sorted(measurement_nodes):
                if node and node.lower() != 'nan':
                    # Создаем чекбокс для узла
                    var = tk.BooleanVar(value=True)  # По умолчанию все выбраны
                    self.node_vars[node] = var
                    ttk.Checkbutton(self.nodes_scrollable_frame, text=node, variable=var).pack(anchor="w", padx=5, pady=2)
                
        except Exception as e:
            error_message = f"Ошибка при обновлении списка узлов измерения: {str(e)}"
            print(error_message)
            messagebox.showerror("Ошибка", error_message)
            
    def remove_empty_columns(self, df):
        """Удаляет пустые столбцы из DataFrame"""
        # Удаляем столбцы, где все значения NaN
        empty_cols = df.columns[df.isna().all()].tolist()
        if empty_cols:
            df = df.drop(columns=empty_cols)
        return df

    def read_excel_file(self, file_path):
        """Читает Excel файл с поддержкой обоих форматов .xls и .xlsx"""
        if file_path.endswith('.xlsx'):
            return pd.read_excel(file_path, engine='openpyxl')
        else:  # для .xls файлов
            return pd.read_excel(file_path, engine='xlrd')
            
    def get_rename_rules(self, filename):
        """Получает правила переименования для конкретного файла"""
        try:
            # Читаем файл с правилами
            rules_file = resource_path("Правила названия столбцов.xlsx")
            if not os.path.exists(rules_file):
                messagebox.showerror("Ошибка", "Файл с правилами названия столбцов не найден!")
                return {}
                
            # Читаем файл с правилами
            rules_df = pd.read_excel(rules_file, engine='openpyxl')
            
            # Ищем правила для данного файла
            rename_dict = {}
            filename = os.path.basename(filename).lower()  # Получаем имя файла без пути
            
            # Получаем список выбранных параметров
            selected_parameters = [param for param, var in self.parameter_vars.items() if var.get()]
            
            # Перебираем все строки в файле правил
            for _, row in rules_df.iterrows():
                # Проверяем, что все необходимые столбцы существуют
                if len(row) >= 5:  # Должно быть как минимум 5 столбцов
                    file_pattern = str(row.iloc[0]).strip().lower()  # Паттерн файла
                    old_name = str(row.iloc[1]).strip()  # Старое название столбца
                    new_name = str(row.iloc[2]).strip()  # Новое название столбца
                    parameter = str(row.iloc[4]).strip()  # Параметр
                    
                    # Если имя файла содержит паттерн и названия не пустые
                    if (pd.notna(file_pattern) and pd.notna(old_name) and pd.notna(new_name) and pd.notna(parameter)
                        and file_pattern and old_name and new_name and parameter
                        and file_pattern in filename
                        and parameter in selected_parameters):  # Проверяем, выбран ли параметр
                        rename_dict[old_name] = new_name
                        print(f"Найдено правило для файла '{filename}': '{old_name}' -> '{new_name}' (Параметр: {parameter})")
            
            return rename_dict
            
        except Exception as e:
            error_message = f"Ошибка при чтении правил переименования: {str(e)}"
            print(error_message)
            messagebox.showerror("Ошибка", error_message)
            return {}
            
    def open_range_editor(self):
        """Открывает окно для редактирования диапазонов значений"""
        editor = tk.Toplevel(self.root)
        editor.title("Редактор диапазонов")
        editor.geometry("600x400")
        
        # Создаем фрейм с прокруткой
        canvas = tk.Canvas(editor)
        scrollbar = ttk.Scrollbar(editor, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Загружаем текущие правила
        try:
            rules_file = resource_path("Правила названия столбцов.xlsx")
            if not os.path.exists(rules_file):
                messagebox.showerror("Ошибка", "Файл с правилами не найден")
                return
            
            # Читаем Excel файл
            df = pd.read_excel(rules_file, engine='openpyxl')
            
            entries = {}
            
            # Группируем по имени параметра (NewName - столбец index 2)
            unique_params = df.iloc[:, 2].unique()
            unique_params = [p for p in unique_params if pd.notna(p)]
            unique_params.sort()
            
            for param_name in unique_params:
                # Получаем текущие значения min/max для этого параметра
                param_rows = df[df.iloc[:, 2] == param_name]
                if param_rows.empty:
                    continue
                    
                current_min = param_rows.iloc[0, 5] if len(param_rows.columns) > 5 else ""
                current_max = param_rows.iloc[0, 6] if len(param_rows.columns) > 6 else ""
                
                if pd.isna(current_min): current_min = ""
                if pd.isna(current_max): current_max = ""
                
                # Пропускаем параметры, у которых не заданы ни мин, ни макс значения
                if current_min == "" and current_max == "":
                    continue
                
                # UI для строки
                frame = ttk.Frame(scrollable_frame)
                frame.pack(fill="x", padx=5, pady=2)
                
                ttk.Label(frame, text=param_name, width=30).pack(side="left")
                
                ttk.Label(frame, text="Min:").pack(side="left")
                min_entry = ttk.Entry(frame, width=10)
                min_entry.insert(0, str(current_min))
                min_entry.pack(side="left", padx=5)
                
                ttk.Label(frame, text="Max:").pack(side="left")
                max_entry = ttk.Entry(frame, width=10)
                max_entry.insert(0, str(current_max))
                max_entry.pack(side="left", padx=5)
                
                entries[param_name] = (min_entry, max_entry)
            
            def save_changes():
                try:
                    for param_name, (min_e, max_e) in entries.items():
                        new_min = min_e.get().strip()
                        new_max = max_e.get().strip()
                        
                        try:
                            val_min = float(new_min) if new_min else None
                        except ValueError:
                            val_min = None
                            
                        try:
                            val_max = float(new_max) if new_max else None
                        except ValueError:
                            val_max = None
                            
                        mask = df.iloc[:, 2] == param_name
                        if val_min is not None:
                            df.loc[mask, df.columns[5]] = val_min
                        else:
                            df.loc[mask, df.columns[5]] = None
                            
                        if val_max is not None:
                            df.loc[mask, df.columns[6]] = val_max
                        else:
                            df.loc[mask, df.columns[6]] = None

                    df.to_excel(rules_file, index=False)
                    messagebox.showinfo("Успех", "Диапазоны обновлены успешно")
                    editor.destroy()
                    
                except Exception as e:
                    messagebox.showerror("Ошибка сохранения", str(e))
            
            ttk.Button(editor, text="Сохранить", command=save_changes).pack(pady=10)
            
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось инициализировать редактор: {e}")

    def update_time_range(self, df):
        """Обновляет диапазон времени на основе данных"""
        try:
            if 'Время' in df.columns:
                time_col = df['Время']
                if pd.api.types.is_datetime64_any_dtype(time_col):
                    start_time = time_col.min()
                    end_time = time_col.max()
                    
                    # Если временной диапазон уже существует, обновляем его
                    if hasattr(self, 'time_range'):
                        current_start, current_end = self.time_range
                        start_time = min(start_time, current_start)
                        end_time = max(end_time, current_end)
                    
                    self.time_range = (start_time, end_time)
                    self.start_date.delete(0, tk.END)
                    self.start_date.insert(0, start_time.strftime('%Y-%m-%d'))
                    self.start_time.delete(0, tk.END)
                    self.start_time.insert(0, start_time.strftime('%H:%M:%S'))
                    
                    self.end_date.delete(0, tk.END)
                    self.end_date.insert(0, end_time.strftime('%Y-%m-%d'))
                    self.end_time.delete(0, tk.END)
                    self.end_time.insert(0, end_time.strftime('%H:%M:%S'))
                    
        except Exception as e:
            print(f"Ошибка при обновлении диапазона времени: {str(e)}")
            
    def set_full_time_range(self):
        """Устанавливает полный диапазон дат из загруженных данных"""
        if hasattr(self, 'time_range'):
            start_time, end_time = self.time_range
            self.start_date.delete(0, tk.END)
            self.start_date.insert(0, start_time.strftime('%Y-%m-%d'))
            self.start_time.delete(0, tk.END)
            self.start_time.insert(0, start_time.strftime('%H:%M:%S'))
            
            self.end_date.delete(0, tk.END)
            self.end_date.insert(0, end_time.strftime('%Y-%m-%d'))
            self.end_time.delete(0, tk.END)
            self.end_time.insert(0, end_time.strftime('%H:%M:%S'))
            
    def clear_dates(self):
        """Очищает поля ввода дат"""
        self.start_date.delete(0, tk.END)
        self.start_time.delete(0, tk.END)
        self.end_date.delete(0, tk.END)
        self.end_time.delete(0, tk.END)
        
    def merge_files(self):
        if not self.files:
            messagebox.showerror("Ошибка", "Пожалуйста, добавьте файлы для объединения")
            return
            
        # Проверяем, выбран ли хотя бы один параметр
        if not any(var.get() for var in self.parameter_vars.values()):
            messagebox.showerror("Ошибка", "Пожалуйста, выберите хотя бы один параметр для объединения")
            return
            
        # Проверяем, выбран ли хотя бы один узел измерения
        if not any(var.get() for var in self.node_vars.values()):
            messagebox.showerror("Ошибка", "Пожалуйста, выберите хотя бы один узел измерения")
            return
            
        try:
            # Проверяем и преобразуем временной интервал
            start_datetime = None
            end_datetime = None
            
            if self.start_date.get() or self.start_time.get():
                try:
                    start_str = f"{self.start_date.get()} {self.start_time.get()}"
                    start_datetime = pd.to_datetime(start_str)
                except Exception as e:
                    messagebox.showerror("Ошибка", f"Неверный формат начальной даты/времени: {str(e)}")
                    return
                    
            if self.end_date.get() or self.end_time.get():
                try:
                    end_str = f"{self.end_date.get()} {self.end_time.get()}"
                    end_datetime = pd.to_datetime(end_str)
                except Exception as e:
                    messagebox.showerror("Ошибка", f"Неверный формат конечной даты/времени: {str(e)}")
                    return
                    
            if start_datetime and end_datetime and start_datetime > end_datetime:
                messagebox.showerror("Ошибка", "Начальная дата/время не может быть позже конечной")
                return
            
            # Читаем файл с правилами для получения соответствия параметров и новых названий
            rules_file = resource_path("Правила названия столбцов.xlsx")
            if not os.path.exists(rules_file):
                messagebox.showerror("Ошибка", "Файл с правилами названия столбцов не найден!")
                return
                
            rules_df = pd.read_excel(rules_file, engine='openpyxl')
            
            # Создаем словарь соответствия параметров и новых названий столбцов
            param_to_columns = {}
            # Создаем словарь соответствия узлов измерения и их столбцов
            node_to_columns = {}
            
            for _, row in rules_df.iterrows():
                if len(row) >= 5:
                    new_name = str(row.iloc[2]).strip()
                    parameter = str(row.iloc[4]).strip()
                    node_name = str(row.iloc[3]).strip()
                    
                    if pd.notna(new_name) and pd.notna(parameter) and pd.notna(node_name) and new_name and parameter and node_name:
                        # Добавляем в словарь параметров
                        if parameter not in param_to_columns:
                            param_to_columns[parameter] = set()
                        param_to_columns[parameter].add(new_name)
                        
                        # Добавляем в словарь узлов
                        if node_name not in node_to_columns:
                            node_to_columns[node_name] = set()
                        node_to_columns[node_name].add(new_name)
            
            # Получаем списки выбранных параметров и узлов
            selected_parameters = [param for param, var in self.parameter_vars.items() if var.get()]
            selected_nodes = [node for node, var in self.node_vars.items() if var.get()]
            
            # Создаем множество допустимых названий столбцов
            allowed_columns = set()
            for param in selected_parameters:
                if param in param_to_columns:
                    allowed_columns.update(param_to_columns[param])
            
            # Фильтруем столбцы по выбранным узлам
            node_allowed_columns = set()
            for node in selected_nodes:
                if node in node_to_columns:
                    node_allowed_columns.update(node_to_columns[node])
            
            # Пересечение множеств для получения только тех столбцов, которые соответствуют и параметрам, и узлам
            allowed_columns = allowed_columns.intersection(node_allowed_columns)
            
            # Чтение всех файлов
            dfs = []
            time_column = None
            
            for i, file in enumerate(self.files):
                print(f"\nОбработка файла: {file}")
                df = self.read_excel_file(file)
                print(f"Столбцы в файле: {list(df.columns)}")
                
                # Удаляем пустые столбцы из каждого файла
                df = self.remove_empty_columns(df)
                
                if i == 0:
                    # В первом файле сохраняем временной столбец
                    time_column = df.iloc[:, 0]  # Предполагаем, что первый столбец - время
                    df = df.iloc[:, 1:]  # Берем все столбцы кроме временного
                else:
                    # В остальных файлах удаляем временной столбец, если он есть
                    if 'Время' in df.columns:
                        df = df.drop(columns=['Время'])
                    # Если первый столбец похож на время (содержит даты или время), удаляем его
                    if pd.api.types.is_datetime64_any_dtype(df.iloc[:, 0]):
                        df = df.iloc[:, 1:]
                
                # Получаем правила переименования для текущего файла
                rename_rules = self.get_rename_rules(file)
                if rename_rules:
                    print(f"Применяем правила переименования для файла {os.path.basename(file)}:")
                    for old_name, new_name in rename_rules.items():
                        print(f"  {old_name} -> {new_name}")
                    # Переименовываем столбцы согласно правилам
                    df = df.rename(columns=rename_rules)
                
                # Оставляем только столбцы, соответствующие выбранным параметрам и узлам
                columns_to_keep = [col for col in df.columns if col in allowed_columns]
                df = df[columns_to_keep]
                
                dfs.append(df)
            
            # Объединение всех датафреймов по столбцам
            merged_df = pd.concat(dfs, axis=1)
            print("\nСтолбцы после объединения:", list(merged_df.columns))
            
            # Добавляем временной столбец в начало
            merged_df.insert(0, 'Время', time_column)
            
            # Фильтруем по временному интервалу, если он указан
            if start_datetime is not None or end_datetime is not None:
                if start_datetime is not None:
                    merged_df = merged_df[merged_df['Время'] >= start_datetime]
                if end_datetime is not None:
                    merged_df = merged_df[merged_df['Время'] <= end_datetime]
                print(f"Применен фильтр по времени: {start_datetime if start_datetime else 'начало'} - {end_datetime if end_datetime else 'конец'}")
            
            # Удаляем пустые столбцы из объединенного датафрейма
            merged_df = self.remove_empty_columns(merged_df)
            
            # Добавляем столбцы со стрелками перед сохранением
            merged_df, param_to_node = add_arrow_columns(merged_df, rules_file)
            
            # Сохранение результата
            output_file = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                title="Сохранить объединенный файл"
            )
            
            if output_file:
                try:
                    print(f"Начинаем сохранение результата в файл: {output_file}")
                    
                    # Сохраняем данные
                    merged_df.to_excel(output_file, index=False, engine='openpyxl')
                    print("Данные успешно сохранены")
                    
                    # Закрываем все открытые файлы Excel
                    import gc
                    gc.collect()
                    
                    # Затем применяем условное форматирование
                    print("Начинаем применять форматирование...")
                    if apply_conditional_formatting(output_file):
                        messagebox.showinfo("Успех", "Файлы успешно объединены и отформатированы!")
                    else:
                        messagebox.showwarning("Предупреждение", 
                            "Файлы объединены, но не удалось применить форматирование.\n"
                            "Проверьте, не открыт ли файл в Excel.")
                except Exception as save_error:
                    error_message = f"Ошибка при сохранении файла: {str(save_error)}"
                    print(error_message)
                    messagebox.showerror("Ошибка", error_message)
                
        except Exception as e:
            error_message = f"Произошла ошибка при объединении файлов: {str(e)}"
            print(error_message)
            messagebox.showerror("Ошибка", error_message)

def setup_logging():
    """Configures logging to a file in the user's home directory."""
    log_dir = os.path.join(os.path.expanduser("~"), ".analytics_ui")
    os.makedirs(log_dir, exist_ok=True)
    log_file = os.path.join(log_dir, "app.log")
    
    logging.basicConfig(
        filename=log_file,
        level=logging.ERROR,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    )
    return log_file

def main():
    try:
        log_file = setup_logging()
        print(f"Logging to: {log_file}")
        print("Starting application...")
        
        root = tk.Tk()
        app = ExcelMerger(root)
        print("Entering main loop...")
        root.mainloop()
    except Exception as e:
        error_msg = f"Critical error: {e}"
        print(error_msg)
        logging.critical(error_msg, exc_info=True)
        import traceback
        traceback.print_exc()
        input("Press Enter to exit...") # Keep window open if run from double-click

if __name__ == "__main__":
    main() 